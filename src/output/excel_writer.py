"""
Excel writer for styled shift roster output.
Produces .xlsx matching the user's existing roster format with:
- Calendar date headers (1-Mar, 2-Mar)
- Day name sub-headers (Mon, Tue)
- Color-coded shift cells
- IRM weekend highlighting
- Summary count rows
- Shift timing legend
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from typing import Dict, List
from pathlib import Path

from src.config.constants import (
    SHIFT_COLORS, SHIFT_FONT_COLORS, SHIFT_ID_TO_LABEL,
    SHIFT_TIMINGS, WORKING_SHIFTS, SHIFT_A, SHIFT_B, SHIFT_C, SHIFT_OG,
)
from src.config.employees import FIXED_EMPLOYEES


# ─── Style Constants ──────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
IRM_HEADER_FILL = PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid")
WEEKEND_HEADER_FILL = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
HOLIDAY_TCS_FILL = PatternFill(start_color="FFFF6666", end_color="FFFF6666", fill_type="solid")
HOLIDAY_CIGNA_FILL = PatternFill(start_color="FF66B2FF", end_color="FF66B2FF", fill_type="solid")

NAME_FONT = Font(name="Calibri", bold=True, size=10)
DATA_FONT = Font(name="Calibri", bold=True, size=10)
LEGEND_FONT = Font(name="Calibri", size=9)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

# Summary row colors
SUMMARY_COLORS = {
    "A": PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid"),
    "B": PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid"),
    "OG": PatternFill(start_color="FFF4B084", end_color="FFF4B084", fill_type="solid"),
    "C": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
}


def _get_shift_fill(label: str) -> PatternFill:
    """Get openpyxl fill for a shift label."""
    color = SHIFT_COLORS.get(label, "FFFFFFFF")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _get_shift_font(label: str) -> Font:
    """Get openpyxl font for a shift label."""
    color = SHIFT_FONT_COLORS.get(label, "FF000000")
    return Font(name="Calibri", bold=True, size=10, color=color)


def write_roster_excel(
    roster_data: Dict,
    calendar_context: Dict,
    output_path: str,
):
    """
    Write the solved roster to a styled Excel file.
    
    Args:
        roster_data: Solver output dict with "roster", "coverage", "stats" keys
        calendar_context: Calendar context from build_calendar_context()
        output_path: Path to output .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar_context['month_name']} {calendar_context['year']}"

    num_days = calendar_context["num_days"]
    day_info = calendar_context["day_info"]
    roster = roster_data["roster"]
    coverage = roster_data["coverage"]

    # Column offset: A=names, B onwards=dates
    date_col_start = 2  # Column B

    # ═══════════════════════════════════════════════════════════════
    #  ROW 1: Title Row
    # ═══════════════════════════════════════════════════════════════
    row = 1
    ws.cell(row=row, column=1, value="Name")
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=row, column=1).border = THIN_BORDER

    for d_idx in range(num_days):
        col = date_col_start + d_idx
        info = day_info[d_idx]
        cell = ws.cell(row=row, column=col, value=info["date_header"])
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")

        # Color based on day type
        if info["is_irm"]:
            cell.fill = IRM_HEADER_FILL
        elif info["is_tcs_holiday"]:
            cell.fill = HOLIDAY_TCS_FILL
        elif info["is_cigna_holiday"]:
            cell.fill = HOLIDAY_CIGNA_FILL
        elif info["is_weekend"]:
            cell.fill = WEEKEND_HEADER_FILL
            cell.font = Font(name="Calibri", bold=True, size=9, color="FF000000")
        else:
            cell.fill = HEADER_FILL

    # ═══════════════════════════════════════════════════════════════
    #  ROW 2: Day Names (Mon, Tue, Wed...)
    # ═══════════════════════════════════════════════════════════════
    row = 2
    ws.cell(row=row, column=1, value="Day")
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=row, column=1).border = THIN_BORDER

    for d_idx in range(num_days):
        col = date_col_start + d_idx
        info = day_info[d_idx]
        cell = ws.cell(row=row, column=col, value=info["day_name"])
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.font = Font(name="Calibri", bold=True, size=9)

        if info["is_irm"]:
            cell.fill = IRM_HEADER_FILL
            cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")
        elif info["is_weekend"]:
            cell.fill = WEEKEND_HEADER_FILL
        elif info["is_tcs_holiday"]:
            cell.fill = HOLIDAY_TCS_FILL
            cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")
        elif info["is_cigna_holiday"]:
            cell.fill = HOLIDAY_CIGNA_FILL
            cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")

    # ═══════════════════════════════════════════════════════════════
    #  ROW 3: IRM Marker
    # ═══════════════════════════════════════════════════════════════
    row = 3
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=1).border = THIN_BORDER

    for d_idx in range(num_days):
        col = date_col_start + d_idx
        info = day_info[d_idx]
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        if info["is_irm"]:
            cell.value = "IRM"
            cell.fill = IRM_HEADER_FILL
            cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")

    # ═══════════════════════════════════════════════════════════════
    #  FIXED EMPLOYEE ROWS (Madhukar, Mukesh, Archith)
    # ═══════════════════════════════════════════════════════════════
    current_row = 4
    for fixed_emp in FIXED_EMPLOYEES:
        ws.cell(row=current_row, column=1, value=fixed_emp.name)
        ws.cell(row=current_row, column=1).font = NAME_FONT
        ws.cell(row=current_row, column=1).alignment = LEFT_ALIGN
        ws.cell(row=current_row, column=1).border = THIN_BORDER

        default_label = SHIFT_ID_TO_LABEL[fixed_emp.fixed_shift]
        for d_idx in range(num_days):
            col = date_col_start + d_idx
            info = day_info[d_idx]

            # Weekend OFF logic for fixed employees (e.g., Archith)
            if fixed_emp.weekend_off and info["is_weekend"]:
                label = "OFF"
            else:
                label = default_label

            cell = ws.cell(row=current_row, column=col, value=label)
            cell.fill = _get_shift_fill(label)
            cell.font = _get_shift_font(label)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        current_row += 1

    # ═══════════════════════════════════════════════════════════════
    #  SOLVER EMPLOYEE ROWS
    # ═══════════════════════════════════════════════════════════════
    for emp_data in roster:
        ws.cell(row=current_row, column=1, value=emp_data["name"])
        ws.cell(row=current_row, column=1).font = NAME_FONT
        ws.cell(row=current_row, column=1).alignment = LEFT_ALIGN
        ws.cell(row=current_row, column=1).border = THIN_BORDER

        for d_idx, shift_label in enumerate(emp_data["shifts"]):
            col = date_col_start + d_idx
            cell = ws.cell(row=current_row, column=col, value=shift_label)
            cell.fill = _get_shift_fill(shift_label)
            cell.font = _get_shift_font(shift_label)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        current_row += 1

    # ═══════════════════════════════════════════════════════════════
    #  EMPTY SEPARATOR ROW
    # ═══════════════════════════════════════════════════════════════
    current_row += 1

    # ═══════════════════════════════════════════════════════════════
    #  SUMMARY COUNT ROWS (A, B, OG, C counts per day)
    # ═══════════════════════════════════════════════════════════════
    summary_shifts = [("A", SHIFT_A), ("B", SHIFT_B), ("OG", SHIFT_OG), ("C", SHIFT_C)]

    for label, shift_id in summary_shifts:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=current_row, column=1).fill = SUMMARY_COLORS.get(label, PatternFill())
        ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=current_row, column=1).border = THIN_BORDER

        for d_idx in range(num_days):
            col = date_col_start + d_idx
            info = day_info[d_idx]
            # Count from solver roster
            count = sum(
                1 for emp_data in roster
                if d_idx < len(emp_data["shifts"]) and emp_data["shifts"][d_idx] == label
            )
            # Add fixed employees (respecting weekend_off logic)
            for fixed_emp in FIXED_EMPLOYEES:
                fixed_label = SHIFT_ID_TO_LABEL[fixed_emp.fixed_shift]
                if fixed_emp.weekend_off and info["is_weekend"]:
                    # Weekend OFF — don't count this employee's shift
                    pass
                elif fixed_label == label:
                    count += 1

            cell = ws.cell(row=current_row, column=col, value=count)
            cell.fill = SUMMARY_COLORS.get(label, PatternFill())
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        current_row += 1

    # ═══════════════════════════════════════════════════════════════
    #  LEGEND SECTION
    # ═══════════════════════════════════════════════════════════════
    current_row += 2

    # Holiday legend
    legend_items = [
        ("TH", "TCS Holiday", HOLIDAY_TCS_FILL),
        ("CH", "Cigna Holiday", HOLIDAY_CIGNA_FILL),
    ]
    for code, desc, fill in legend_items:
        ws.cell(row=current_row, column=1, value=code)
        ws.cell(row=current_row, column=1).fill = fill
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", bold=True, size=9, color="FFFFFFFF")
        ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=current_row, column=2, value=desc)
        ws.cell(row=current_row, column=2).font = LEGEND_FONT
        current_row += 1

    current_row += 1

    # Shift timing legend
    for shift_label, timings in SHIFT_TIMINGS.items():
        ws.cell(row=current_row, column=1, value=shift_label)
        ws.cell(row=current_row, column=1).fill = _get_shift_fill(shift_label)
        ws.cell(row=current_row, column=1).font = _get_shift_font(shift_label)
        ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=current_row, column=1).border = THIN_BORDER

        timing_str = f"{shift_label} Shift ({timings['IST']} IST) = ({timings['EST']} EST)"
        ws.cell(row=current_row, column=2, value=timing_str)
        ws.cell(row=current_row, column=2).font = LEGEND_FONT
        current_row += 1

    # ═══════════════════════════════════════════════════════════════
    #  COLUMN WIDTHS
    # ═══════════════════════════════════════════════════════════════
    ws.column_dimensions["A"].width = 16  # Name column
    for d_idx in range(num_days):
        col_letter = get_column_letter(date_col_start + d_idx)
        ws.column_dimensions[col_letter].width = 7  # Date columns

    # ═══════════════════════════════════════════════════════════════
    #  FREEZE PANES (Freeze name column + header rows)
    # ═══════════════════════════════════════════════════════════════
    ws.freeze_panes = "B4"

    # ═══════════════════════════════════════════════════════════════
    #  SAVE
    # ═══════════════════════════════════════════════════════════════
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"    📊 Excel roster saved to: {output}")
